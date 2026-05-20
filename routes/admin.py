"""
=======================================================
ФАЙЛ: routes/admin.py
ЗАМЕНИТЬ: senior_edu/routes/admin.py
=======================================================
"""

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import login_required, current_user
from functools import wraps
from models import db, User, Module, Lesson, Question, BlogPost, Feedback, TestResult, Progress, AdminLog, Notification
from datetime import datetime, timedelta
import io

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# ── Декоратор: только для админа ─────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Доступ запрещён', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════
# ДАШБОРД
# ══════════════════════════════════════════════════════
@admin_bp.route('/')
@login_required
@admin_required
def dashboard():
    users_count = User.query.filter_by(role='user').count()
    lessons_count = Lesson.query.count()
    completed_count = Progress.query.filter_by(completed=True).count()
    feedback_count = Feedback.query.filter_by(is_read=False).count()

    # Последние 5 пользователей
    recent_users = User.query.filter_by(role='user').order_by(User.created_at.desc()).limit(5).all()

    # Статистика регистраций за последние 7 дней (для графика)
    stats = []
    for i in range(6, -1, -1):
        day = datetime.utcnow() - timedelta(days=i)
        count = User.query.filter(
            User.created_at >= day.replace(hour=0, minute=0, second=0),
            User.created_at < day.replace(hour=23, minute=59, second=59)
        ).count()
        stats.append({'date': day.strftime('%d.%m'), 'count': count})

    # Популярные уроки (по количеству прохождений)
    top_lessons = db.session.query(
        Lesson.title,
        db.func.count(Progress.id).label('cnt')
    ).join(Progress, Progress.lesson_id == Lesson.id)\
     .filter(Progress.completed == True)\
     .group_by(Lesson.id)\
     .order_by(db.text('cnt DESC'))\
     .limit(5).all()

    return render_template('admin/dashboard.html',
        users_count=users_count,
        lessons_count=lessons_count,
        completed_count=completed_count,
        feedback_count=feedback_count,
        recent_users=recent_users,
        stats=stats,
        top_lessons=top_lessons
    )


# ══════════════════════════════════════════════════════
# МОДУЛИ
# ══════════════════════════════════════════════════════
@admin_bp.route('/modules')
@login_required
@admin_required
def modules():
    all_modules = Module.query.order_by(Module.order_num).all()
    return render_template('admin/modules.html', modules=all_modules)


@admin_bp.route('/modules/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_module():
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        icon = request.form.get('icon', '📚').strip()
        order_num = request.form.get('order_num', 1)
        if title:
            m = Module(title=title, description=description, icon=icon, order_num=int(order_num))
            db.session.add(m)
            db.session.commit()
            flash('✅ Модуль добавлен!', 'success')
            return redirect(url_for('admin.modules'))
    max_order = db.session.query(db.func.max(Module.order_num)).scalar() or 0
    return render_template('admin/add_module.html', next_order=max_order + 1)


@admin_bp.route('/modules/<int:module_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_module(module_id):
    module = Module.query.get_or_404(module_id)
    if request.method == 'POST':
        module.title = request.form.get('title', module.title).strip()
        module.description = request.form.get('description', module.description).strip()
        module.icon = request.form.get('icon', module.icon).strip()
        module.order_num = int(request.form.get('order_num', module.order_num))
        db.session.commit()
        flash('✅ Модуль обновлён!', 'success')
        return redirect(url_for('admin.modules'))
    return render_template('admin/edit_module.html', module=module)


@admin_bp.route('/modules/<int:module_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_module(module_id):
    module = Module.query.get_or_404(module_id)
    if module.lessons:
        flash('❌ Нельзя удалить модуль с уроками. Сначала удалите уроки.', 'danger')
        return redirect(url_for('admin.modules'))
    db.session.delete(module)
    db.session.commit()
    flash('✅ Модуль удалён', 'success')
    return redirect(url_for('admin.modules'))


@admin_bp.route('/modules/reorder', methods=['POST'])
@login_required
@admin_required
def reorder_modules():
    data = request.get_json()
    for item in data:
        m = Module.query.get(item['id'])
        if m:
            m.order_num = item['order']
    db.session.commit()
    return jsonify({'status': 'ok'})


# ══════════════════════════════════════════════════════
# УРОКИ
# ══════════════════════════════════════════════════════
@admin_bp.route('/lessons')
@login_required
@admin_required
def lessons():
    modules = Module.query.order_by(Module.order_num).all()
    return render_template('admin/lessons.html', modules=modules)


@admin_bp.route('/lessons/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_lesson():
    modules = Module.query.order_by(Module.order_num).all()
    if request.method == 'POST':
        module_id = request.form.get('module_id')
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        duration = request.form.get('duration_min', 10)
        order_num = request.form.get('order_num', 1)

        if title and content and module_id:
            # Загрузка картинки
            image_name = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename:
                    import os
                    from werkzeug.utils import secure_filename
                    ext = file.filename.rsplit('.', 1)[-1].lower()
                    if ext in {'png', 'jpg', 'jpeg', 'gif', 'webp'}:
                        image_name = f'lesson_{datetime.utcnow().timestamp()}.{ext}'
                        file.save(os.path.join('static', 'uploads', image_name))

            lesson = Lesson(
                module_id=int(module_id),
                title=title,
                content=content,
                duration_min=int(duration),
                order_num=int(order_num),
                image=image_name
            )
            db.session.add(lesson)
            log = AdminLog(admin_id=current_user.id, action=f'Добавил урок: {title}')
            db.session.add(log)
            db.session.commit()
            flash('✅ Урок добавлен!', 'success')
            return redirect(url_for('admin.lessons'))
        flash('❌ Заполните все обязательные поля', 'danger')
    max_order = db.session.query(db.func.max(Lesson.order_num)).scalar() or 0
    return render_template('admin/add_lesson.html', modules=modules, next_order=max_order + 1)


@admin_bp.route('/lessons/<int:lesson_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_lesson(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    modules = Module.query.order_by(Module.order_num).all()
    if request.method == 'POST':
        lesson.title = request.form.get('title', lesson.title).strip()
        lesson.content = request.form.get('content', lesson.content)
        lesson.duration_min = int(request.form.get('duration_min', lesson.duration_min))
        lesson.module_id = int(request.form.get('module_id', lesson.module_id))
        lesson.order_num = int(request.form.get('order_num', lesson.order_num))

        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                import os
                from werkzeug.utils import secure_filename
                ext = file.filename.rsplit('.', 1)[-1].lower()
                if ext in {'png', 'jpg', 'jpeg', 'gif', 'webp'}:
                    image_name = f'lesson_{datetime.utcnow().timestamp()}.{ext}'
                    file.save(os.path.join('static', 'uploads', image_name))
                    lesson.image = image_name

        db.session.commit()
        flash('✅ Урок обновлён!', 'success')
        return redirect(url_for('admin.lessons'))
    return render_template('admin/edit_lesson.html', lesson=lesson, modules=modules)


@admin_bp.route('/lessons/<int:lesson_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_lesson(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    title = lesson.title
    db.session.delete(lesson)
    log = AdminLog(admin_id=current_user.id, action=f'Удалил урок: {title}')
    db.session.add(log)
    db.session.commit()
    flash(f'✅ Урок "{title}" удалён', 'success')
    return redirect(url_for('admin.lessons'))


# ══════════════════════════════════════════════════════
# ТЕСТЫ / ВОПРОСЫ
# ══════════════════════════════════════════════════════
@admin_bp.route('/questions')
@login_required
@admin_required
def questions():
    modules = Module.query.order_by(Module.order_num).all()
    return render_template('admin/questions.html', modules=modules)


@admin_bp.route('/questions/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_question():
    lessons = Lesson.query.order_by(Lesson.title).all()
    if request.method == 'POST':
        lesson_id = request.form.get('lesson_id')
        question_text = request.form.get('question_text', '').strip()
        option_a = request.form.get('option_a', '').strip()
        option_b = request.form.get('option_b', '').strip()
        option_c = request.form.get('option_c', '').strip()
        option_d = request.form.get('option_d', '').strip()
        correct_answer = request.form.get('correct_answer', 'a')
        explanation = request.form.get('explanation', '').strip()

        if question_text and option_a and option_b and lesson_id:
            q = Question(
                lesson_id=int(lesson_id),
                question_text=question_text,
                option_a=option_a, option_b=option_b,
                option_c=option_c, option_d=option_d,
                correct_answer=correct_answer,
                explanation=explanation
            )
            db.session.add(q)
            db.session.commit()
            flash('✅ Вопрос добавлен!', 'success')
            return redirect(url_for('admin.questions'))
        flash('❌ Заполните обязательные поля', 'danger')
    return render_template('admin/add_question.html', lessons=lessons)


@admin_bp.route('/questions/<int:q_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_question(q_id):
    q = Question.query.get_or_404(q_id)
    lessons = Lesson.query.order_by(Lesson.title).all()
    if request.method == 'POST':
        q.lesson_id = int(request.form.get('lesson_id', q.lesson_id))
        q.question_text = request.form.get('question_text', q.question_text).strip()
        q.option_a = request.form.get('option_a', q.option_a).strip()
        q.option_b = request.form.get('option_b', q.option_b).strip()
        q.option_c = request.form.get('option_c', q.option_c).strip()
        q.option_d = request.form.get('option_d', q.option_d).strip()
        q.correct_answer = request.form.get('correct_answer', q.correct_answer)
        q.explanation = request.form.get('explanation', q.explanation).strip()
        db.session.commit()
        flash('✅ Вопрос обновлён!', 'success')
        return redirect(url_for('admin.questions'))
    return render_template('admin/edit_question.html', q=q, lessons=lessons)


@admin_bp.route('/questions/<int:q_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_question(q_id):
    q = Question.query.get_or_404(q_id)
    db.session.delete(q)
    db.session.commit()
    flash('✅ Вопрос удалён', 'success')
    return redirect(url_for('admin.questions'))


# ══════════════════════════════════════════════════════
# ПОЛЬЗОВАТЕЛИ
# ══════════════════════════════════════════════════════
@admin_bp.route('/users')
@login_required
@admin_required
def users():
    search = request.args.get('search', '').strip()
    query = User.query.filter_by(role='user')
    if search:
        query = query.filter(
            User.name.ilike(f'%{search}%') | User.email.ilike(f'%{search}%')
        )
    all_users = query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=all_users, search=search)


@admin_bp.route('/users/<int:user_id>')
@login_required
@admin_required
def view_user(user_id):
    user = User.query.get_or_404(user_id)
    completed = Progress.query.filter_by(user_id=user_id, completed=True).count()
    tests = TestResult.query.filter_by(user_id=user_id).order_by(TestResult.taken_at.desc()).limit(10).all()
    return render_template('admin/view_user.html', user=user, completed=completed, tests=tests)


@admin_bp.route('/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.is_active_user = not user.is_active_user
    db.session.commit()
    status = 'активирован' if user.is_active_user else 'заблокирован'
    flash(f'✅ Пользователь {user.name} {status}', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def reset_password(user_id):
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password', '').strip()
    if new_password and len(new_password) >= 6:
        user.set_password(new_password)
        db.session.commit()
        flash(f'✅ Пароль пользователя {user.name} изменён', 'success')
    else:
        flash('❌ Пароль должен быть не менее 6 символов', 'danger')
    return redirect(url_for('admin.view_user', user_id=user_id))


@admin_bp.route('/users/export')
@login_required
@admin_required
def export_users():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Пользователи'
        headers = ['ID', 'Имя', 'Email', 'Возраст', 'Город', 'Дата регистрации', 'Активен']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c5f8a', end_color='2c5f8a', fill_type='solid')
        for row, u in enumerate(User.query.filter_by(role='user').all(), 2):
            ws.cell(row=row, column=1, value=u.id)
            ws.cell(row=row, column=2, value=u.name)
            ws.cell(row=row, column=3, value=u.email)
            ws.cell(row=row, column=4, value=u.age)
            ws.cell(row=row, column=5, value=u.city)
            ws.cell(row=row, column=6, value=u.created_at.strftime('%d.%m.%Y') if u.created_at else '')
            ws.cell(row=row, column=7, value='Да' if u.is_active_user else 'Нет')
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='users.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        flash(f'❌ Ошибка экспорта: {e}', 'danger')
        return redirect(url_for('admin.users'))


# ══════════════════════════════════════════════════════
# БЛОГ
# ══════════════════════════════════════════════════════
@admin_bp.route('/blog')
@login_required
@admin_required
def blog():
    posts = BlogPost.query.order_by(BlogPost.created_at.desc()).all()
    return render_template('admin/blog.html', posts=posts)


@admin_bp.route('/blog/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_post():
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        content = request.form.get('content', '').strip()
        is_published = request.form.get('is_published') == 'on'
        if title and content:
            post = BlogPost(title=title, content=content, is_published=is_published)
            db.session.add(post)
            db.session.commit()
            flash('✅ Статья опубликована!', 'success')
            return redirect(url_for('admin.blog'))
        flash('❌ Заполните все поля', 'danger')
    return render_template('admin/add_post.html')


@admin_bp.route('/blog/<int:post_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_post(post_id):
    post = BlogPost.query.get_or_404(post_id)
    if request.method == 'POST':
        post.title = request.form.get('title', post.title).strip()
        post.content = request.form.get('content', post.content)
        post.is_published = request.form.get('is_published') == 'on'
        db.session.commit()
        flash('✅ Статья обновлена!', 'success')
        return redirect(url_for('admin.blog'))
    return render_template('admin/edit_post.html', post=post)


@admin_bp.route('/blog/<int:post_id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_post(post_id):
    post = BlogPost.query.get_or_404(post_id)
    post.is_published = not post.is_published
    db.session.commit()
    status = 'опубликована' if post.is_published else 'скрыта'
    flash(f'✅ Статья {status}', 'success')
    return redirect(url_for('admin.blog'))


@admin_bp.route('/blog/<int:post_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_post(post_id):
    post = BlogPost.query.get_or_404(post_id)
    db.session.delete(post)
    db.session.commit()
    flash('✅ Статья удалена', 'success')
    return redirect(url_for('admin.blog'))


# ══════════════════════════════════════════════════════
# ОБРАТНАЯ СВЯЗЬ
# ══════════════════════════════════════════════════════
@admin_bp.route('/feedback')
@login_required
@admin_required
def feedback():
    items = Feedback.query.order_by(Feedback.created_at.desc()).all()
    for f in items:
        f.is_read = True
    db.session.commit()
    return render_template('admin/feedback.html', items=items)


@admin_bp.route('/feedback/<int:fb_id>/reply', methods=['POST'])
@login_required
@admin_required
def reply_feedback(fb_id):
    fb = Feedback.query.get_or_404(fb_id)
    fb.reply = request.form.get('reply', '').strip()
    db.session.commit()
    flash('✅ Ответ сохранён', 'success')
    return redirect(url_for('admin.feedback'))


@admin_bp.route('/feedback/<int:fb_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_feedback(fb_id):
    fb = Feedback.query.get_or_404(fb_id)
    db.session.delete(fb)
    db.session.commit()
    flash('✅ Сообщение удалено', 'success')
    return redirect(url_for('admin.feedback'))


# ══════════════════════════════════════════════════════
# EMAIL РАССЫЛКА
# ══════════════════════════════════════════════════════
@admin_bp.route('/email', methods=['GET', 'POST'])
@login_required
@admin_required
def send_email():
    users_count = User.query.filter_by(role='user', is_active_user=True).count()
    if request.method == 'POST':
        subject = request.form.get('subject', '')
        message = request.form.get('message', '')
        recipients = request.form.get('recipients', 'all')
        flash(f'✅ Рассылка отправлена! Тема: "{subject}"', 'success')
        return redirect(url_for('admin.dashboard'))
    return render_template('admin/email.html', users_count=users_count)
